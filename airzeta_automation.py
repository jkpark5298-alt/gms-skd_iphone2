import os
import re
import tempfile
import xml.etree.ElementTree as ET
from datetime import datetime
from tkinter import Tk, filedialog, messagebox

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()

def normalize_header(value):
    text = normalize_text(value).upper()
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text

def detect_file_kind(input_path):
    ext = os.path.splitext(input_path)[1].lower()
    try:
        with open(input_path, "rb") as f:
            head = f.read(512)
    except Exception:
        head = b""
    head_strip = head.lstrip()
    if head.startswith(b"PK"):
        return "xlsx_zip"
    if head.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
        return "xls_binary"
    if head_strip.startswith(b"<?xml") or head_strip.startswith(b"<"):
        return "xml_text"
    if ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        return "xlsx_zip"
    return "unknown"

def parse_xml_to_xlsx_temp(xml_path):
    temp_fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(temp_fd)
    try:
        df = pd.read_xml(xml_path)
        df.to_excel(temp_path, index=False)
        return temp_path
    except Exception:
        try:
            df_list = pd.read_html(xml_path)
            if df_list:
                df_list[0].to_excel(temp_path, index=False)
                return temp_path
        except Exception:
            pass
    return None

def get_header_map(ws, row_idx):
    header_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=row_idx, column=col).value
        if val:
            header_map[normalize_header(val)] = col
    return header_map

def find_column(header_map, keywords):
    for k in keywords:
        norm_k = normalize_header(k)
        if norm_k in header_map:
            return header_map[norm_k]
    return None

def move_column_to_front(ws, col_idx):
    ws.insert_cols(1)
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=1).value = ws.cell(row=row, column=col_idx + 1).value
    ws.delete_cols(col_idx + 1)

def insert_title_row(ws, title_text):
    ws.insert_rows(1)
    ws.cell(row=1, column=1).value = title_text
    ws.cell(row=1, column=1).font = Font(size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

def write_footer_time(ws, time_text):
    last_row = ws.max_row + 1
    ws.cell(row=last_row, column=ws.max_column).value = time_text
    ws.cell(row=last_row, column=ws.max_column).alignment = Alignment(horizontal="right")

def set_all_font_black(ws):
    from openpyxl.styles import Color
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(color="000000")

def auto_fit_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    length = len(str(cell.value).encode('utf-8'))
                    if length > max_length:
                        max_length = length
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

def parse_sort_time(val):
    if not val: return "99:99"
    s = str(val).strip()
    match = re.search(r"(\d{1,2}:\d{2})", s)
    if match:
        return match.group(1)
    return "99:99"

def process_single_file(input_path, output_dir):
    temp_file_to_delete = None
    kind = detect_file_kind(input_path)
    
    if kind == "xml_text":
        temp_file_to_delete = parse_xml_to_xlsx_temp(input_path)
        if not temp_file_to_delete:
            return None
        current_path = temp_file_to_delete
    else:
        current_path = input_path

    try:
        if kind == "xls_binary":
            df = pd.read_excel(current_path, engine='xlrd')
            temp_fd, current_path = tempfile.mkstemp(suffix=".xlsx")
            os.close(temp_fd)
            temp_file_to_delete = current_path
            df.to_excel(current_path, index=False)

        wb = load_workbook(current_path)
        ws = wb.active

        # Find Header
        header_row = 1
        for r in range(1, 6):
            vals = [ws.cell(row=r, column=c).value for c in range(1, 6)]
            if any(vals):
                header_row = r
                break
        
        header_map = get_header_map(ws, header_row)
        col_etd_eta = find_column(header_map, ["ETD/ETA", "ETD", "ETA"])
        col_ro_ld = find_column(header_map, ["R/O L/D", "R/O", "L/D"])
        col_to_ri = find_column(header_map, ["T/O R/I", "T/O", "R/I"])
        col_std = find_column(header_map, ["STD"])
        col_sta = find_column(header_map, ["STA"])

        # Row Deletion & Data Collection
        data_rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            if not any(row_vals): continue
            
            # Condition: ETD/ETA, R/O L/D, T/O R/I all exist
            val1 = ws.cell(row=r, column=col_etd_eta).value if col_etd_eta else None
            val2 = ws.cell(row=r, column=col_ro_ld).value if col_ro_ld else None
            val3 = ws.cell(row=r, column=col_to_ri).value if col_to_ri else None
            
            if val1 and val2 and val3:
                continue # delete
            data_rows.append(row_vals)

        # Sort by ETD/ETA
        if col_etd_eta:
            data_rows.sort(key=lambda x: parse_sort_time(x[col_etd_eta-1]))

        # Clear and Re-write
        ws.delete_rows(header_row + 1, ws.max_row)
        for r_idx, row_data in enumerate(data_rows, start=header_row + 1):
            for c_idx, val in enumerate(row_data, start=1):
                ws.cell(row=r_idx, column=c_idx).value = val

        # Delete STD/STA columns
        cols_to_del = sorted([c for c in [col_std, col_sta] if c], reverse=True)
        for c in cols_to_del:
            ws.delete_cols(c)

        # Refresh header map and move Parking
        header_map = get_header_map(ws, header_row)
        col_parking = find_column(header_map, ["주기장"])
        if col_parking:
            move_column_to_front(ws, col_parking)

        # Title & Footer
        today_display = datetime.now().strftime("%Y-%m-%d")
        insert_title_row(ws, f"에어제타 ({today_display})")
        write_footer_time(ws, "작성시간: " + datetime.now().strftime("%Y-%m-%d %H:%M"))

        set_all_font_black(ws)
        auto_fit_width(ws)

        # Multi-file support: use original filename in output
        base_name = os.path.splitext(os.path.basename(input_path))[0]
        out_name = f"에어제타_{base_name}_{datetime.now().strftime('%H%M%S')}.xlsx"
        out_path = os.path.join(output_dir, out_name)
        wb.save(out_path)
        return out_path
    finally:
        if temp_file_to_delete and os.path.exists(temp_file_to_delete):
            os.remove(temp_file_to_delete)

def main():
    root = Tk()
    root.withdraw()

    input_paths = filedialog.askopenfilenames(
        title="원본 파일들을 선택하세요 (복수 선택 가능)",
        filetypes=[("지원 파일", "*.xlsx *.xlsm *.xltx *.xltm *.xls *.xml *.xma"), ("All files", "*.*")]
    )
    
    if not input_paths:
        return

    output_dir = filedialog.askdirectory(title="저장 폴더 선택")
    if not output_dir:
        return

    # 1. Sort files by Creation/Modification time
    # User requested '생성(수정) 시간' -> os.path.getmtime is standard for modification.
    file_list = list(input_paths)
    file_list.sort(key=lambda p: os.path.getmtime(p))

    success_count = 0
    for p in file_list:
        try:
            res = process_single_file(p, output_dir)
            if res:
                success_count += 1
        except Exception as e:
            print(f"Error processing {p}: {e}")

    messagebox.showinfo("완료", f"총 {len(file_list)}개 중 {success_count}개 파일 처리가 완료되었습니다.")

if __name__ == "__main__":
    main()
